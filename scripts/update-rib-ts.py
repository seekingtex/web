import openpyxl, re, os

BASE = r'C:\seekingtex RIB boat'
TS_PATH = r'C:\aweb\src\data\products.ts'

model_warranty = {}
for model_id in sorted(os.listdir(BASE)):
    xlsx_path = os.path.join(BASE, model_id, model_id + '.xlsx')
    if not os.path.exists(xlsx_path):
        continue
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        key = str(row[0]).strip().rstrip(':').strip() if row[0] else ''
        if key == 'Warranty' and row[1]:
            model_warranty[model_id] = str(row[1]).strip()

with open(TS_PATH, 'r', encoding='utf-8') as f:
    content = f.read()

# Fix air_chambers: ': 3' -> '3' bug
content = content.replace("air_chambers: ': 3'", "air_chambers: '3'")

# Find each model by id and add warranty
for model_id, warranty in sorted(model_warranty.items()):
    slug = 'rib-' + model_id.lower().replace('rib', '')
    id_pattern = f"    id: '{slug}',"
    idx = content.find(id_pattern)
    if idx == -1:
        print(f'WARN: {model_id} ({slug}) not found')
        continue

    ai_specs_start = content.find('ai_specs:', idx)
    if ai_specs_start == -1:
        print(f'WARN: {model_id} - ai_specs not found')
        continue

    ai_comp = content.find('ai_comparison:', ai_specs_start)
    if ai_comp == -1:
        print(f'WARN: {model_id} - ai_comparison not found')
        continue

    block = content[ai_specs_start:ai_comp]
    if 'warranty' in block:
        print(f'SKIP {model_id}: warranty already in ai_specs')
        continue

    # Find the last newline before ai_comparison for insertion
    # We want to add warranty before the closing }, of ai_specs
    last_newline = content.rfind('\n', ai_specs_start, ai_comp)
    if last_newline == -1:
        print(f'WARN: {model_id} - cannot find insertion point')
        continue

    # Find the "}," that closes ai_specs
    close_brace = content.rfind('\n    },', ai_specs_start, ai_comp)
    if close_brace == -1:
        print(f'WARN: {model_id} - closing brace not found')
        continue

    content = content[:close_brace] + f"\n      warranty: '{warranty}'," + content[close_brace:]
    print(f'ADDED warranty={warranty} for {model_id}')

with open(TS_PATH, 'w', encoding='utf-8') as f:
    f.write(content)
print('Done.')
