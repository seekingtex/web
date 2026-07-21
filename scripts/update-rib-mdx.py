import openpyxl, os, re

BASE = r'C:\seekingtex RIB boat'
OUT = r'C:\aweb\src\content\products'

MODEL_NAMES = {
    'RIB250': 'RIB250', 'RIB270': 'RIB270', 'RIB300': 'RIB300',
    'RIB330A': 'RIB330A', 'RIB330B': 'RIB330B', 'RIB330C': 'RIB330C',
    'RIB360A': 'RIB360A', 'RIB360B': 'RIB360B',
    'RIB390A': 'RIB390A', 'RIB390AL': 'RIB390AL', 'RIB390B': 'RIB390B', 'RIB390BL': 'RIB390BL',
    'RIB430A': 'RIB430A', 'RIB430B': 'RIB430B', 'RIB430C': 'RIB430C',
    'RIB480A': 'RIB480A', 'RIB480B': 'RIB480B', 'RIB480C': 'RIB480C', 'RIB480D': 'RIB480D', 'RIB480W': 'RIB480W',
    'RIB500': 'RIB500',
    'RIB520A': 'RIB520A', 'RIB520B': 'RIB520B', 'RIB520C': 'RIB520C', 'RIB520E': 'RIB520E', 'RIB520W': 'RIB520W',
    'RIB550A': 'RIB550A', 'RIB550B': 'RIB550B',
    'RIB580A': 'RIB580A', 'RIB580B': 'RIB580B',
    'RIB600': 'RIB600',
    'RIB680A': 'RIB680A', 'RIB680B': 'RIB680B',
    'RIB700': 'RIB700',
    'RIB760': 'RIB760',
}

def parse_length(cm_str):
    cm_str = str(cm_str).strip().replace('cm', '').strip()
    try:
        cm = float(cm_str)
        return cm / 100, f'{cm:.0f}cm'
    except:
        return None, cm_str

def parse_weight(kg_str):
    kg_str = str(kg_str).strip().replace('kgs', '').replace('kg', '').strip()
    return kg_str + 'kgs'

def get_tags(model_id, length_m, engine_str):
    tags = ['rib', 'commercial']
    if length_m is not None:
        if length_m < 4:
            tags.append('beginner')
        elif length_m < 6:
            tags.append('intermediate')
        else:
            tags.append('advanced')
    else:
        tags.append('intermediate')
    return tags

def get_use_case(length_m):
    if length_m is None:
        return 'lake, coastal'
    if length_m < 3.5:
        return 'lake, coastal'
    elif length_m < 5:
        return 'coastal, sea'
    else:
        return 'coastal, offshore'

def get_model_slug(model_id):
    return 'rib-' + model_id.lower().replace('rib', '')

def read_xlsx(model_id):
    xlsx_path = os.path.join(BASE, model_id, model_id + '.xlsx')
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    specs = {}
    std_accessories = []
    opt_accessories = []
    section = 'specs'

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
        a, b, c, d = (row + (None,) * 4)[:4]

        # Detect sections
        if a and isinstance(a, str):
            a_clean = a.strip().rstrip(':').strip()
            if a_clean in ('Standard Accessories', 'Standard Features'):
                section = 'std'
                continue
            elif a_clean in ('Optional Accessories', 'Optional Features/Accessories'):
                section = 'opt'
                continue

        if section == 'specs':
            if a and isinstance(a, str):
                key = a.strip().rstrip(':').strip()
                val = b
                if key.startswith('Model:'):
                    continue
                specs[key] = val
        elif section == 'std':
            if a:
                txt = str(a).strip().lstrip('*').strip()
                if txt and not txt.startswith('Standard'):
                    std_accessories.append(txt)
        elif section == 'opt':
            if a:
                txt = str(a).strip().lstrip('*').strip()
                if txt and not txt.startswith('Optional'):
                    opt_accessories.append(txt)

        # For models with C/D column accessories (RIB390A style)
        if c and section == 'specs':
            txt = str(c).strip()
            if txt and not txt.startswith('Standard'):
                std_accessories.append(txt)
        if d and section == 'specs':
            txt = str(d).strip()
            if txt and not txt.startswith('Optional'):
                opt_accessories.append(txt)

    return specs, std_accessories, opt_accessories

def generate_frontmatter(model_id, specs, std_acc, opt_acc):
    title = model_id
    slug = get_model_slug(model_id)
    sku = 'RIB-' + model_id.replace('RIB', '')

    total_len = specs.get('Total length', specs.get('Total length:', ''))
    total_width = specs.get('Total width', specs.get('Total width:', ''))
    tube_dia = specs.get('Tube diameter', specs.get('Tube diameter:', ''))
    net_weight = specs.get('Net weight', specs.get('Net weight:', ''))
    max_load = specs.get('Max load', specs.get('Max load:', ''))
    max_pass = specs.get('Max Passenger', specs.get('Max Passenger:', ''))
    air_chamber = specs.get('Air chamber', specs.get('Air chamber:', ''))
    max_engine = specs.get('Max Engine', specs.get('Max Engine:', ''))
    warranty = specs.get('Warranty', specs.get('Warranty: ', ''))

    # Clean leading colon from xlsx values (e.g. ': 3' 鈫?'3')
    def clean_val(v):
        s = str(v).strip()
        if s.startswith(': '):
            s = s[2:]
        return s

    length_m, length_cm = parse_length(clean_val(total_len))
    beam_cm = clean_val(total_width).replace('cm', '').strip()
    tube_cm = clean_val(tube_dia).replace('cm', '').strip()
    net_weight = clean_val(net_weight)
    max_load = clean_val(max_load)
    air_chamber = clean_val(air_chamber)
    max_engine = clean_val(max_engine)
    warranty = clean_val(warranty)
    use_case = get_use_case(length_m)
    tags = get_tags(model_id, length_m, str(max_engine))

    summary = f'High-performance {str(total_len).strip()} rigid-hull inflatable boat ({title}) for transport, fishing.'
    description = f'A {str(total_len).strip()} RIB ({title}) designed for {use_case} conditions, supporting engines up to {str(max_engine).strip()}.'

    gallery_img = f'/images/seekingtex/products/rib/seekingtex-{slug}.jpg'

    # Build specs list
    spec_items = []
    if length_m:
        spec_items.append(f'  - label: Overall length\n    value: {length_m:.1f} m')
    if beam_cm:
        spec_items.append(f'  - label: Beam\n    value: {beam_cm}cm')
    if tube_cm:
        spec_items.append(f'  - label: Tube diameter\n    value: {tube_cm}cm')
    if air_chamber:
        spec_items.append(f'  - label: Air chambers\n    value: {air_chamber}')
    if net_weight:
        spec_items.append(f'  - label: Net weight\n    value: {parse_weight(net_weight)}')
    if max_load:
        spec_items.append(f'  - label: Max load\n    value: {parse_weight(max_load)}')
    if max_pass:
        spec_items.append(f'  - label: Capacity\n    value: {max_pass} persons')
    if max_engine:
        spec_items.append(f'  - label: Max engine\n    value: {str(max_engine).strip()}')
    if warranty:
        spec_items.append(f'  - label: Warranty\n    value: {str(warranty).strip()}')

    specs_yaml = '\n'.join(spec_items)

    # Gallery
    gallery_lines = []
    found_urls = set()
    
    def add_gallery(url, title, view_num):
        gallery_lines.append(f'  - url: {url}\n    alt: "{title} 鈥?view {view_num}"')
        found_urls.add(url)
    
    add_gallery(gallery_img, title, 1)

    # Check for additional gallery images (various naming patterns)
    img_dir = r'C:\aweb\public\images\seekingtex\products\rib'
    base_name = f'seekingtex-{slug}'
    model_lower = model_id.lower()
    patterns = [
        f'{base_name}-{{n}}',
        f'{base_name}-{model_lower}-{{n}}',
        f'{model_lower}-{{n}}',
    ]
    found_urls = set()
    found_urls.add(gallery_img)
    for n in range(1, 10):
        for pat_tmpl in patterns:
            for ext in ['.jpg', '.webp']:
                fname = pat_tmpl.replace('{n}', str(n)) + ext
                full_path = os.path.join(img_dir, fname)
                if os.path.exists(full_path):
                    url = f'/images/seekingtex/products/rib/{fname}'
                    if url not in found_urls:
                        found_urls.add(url)
                        add_gallery(url, title, len(found_urls))
                        break

    gallery_yaml = '\n'.join(gallery_lines)

    # Tags
    tag_yaml = '\n'.join(f'  - {t}' for t in tags)

    # Metadata
    meta_title = f'{title} 鈥?{summary}'
    meta_desc = description

    frontmatter = f'''---
publishDate: 2025-01-15
draft: false
title: {title}
sku: {sku}
summary: {summary}
description: {description}
image: {gallery_img}
gallery:
{gallery_yaml}
category: rib
tags:
{tag_yaml}
specs:
{specs_yaml}
inStock: true
featured: false
metadata:
  title: {meta_title}
  description: {meta_desc}
---
'''
    return frontmatter, slug

def generate_body(model_id, specs, std_acc, opt_acc, slug, length_m):
    total_len = specs.get('Total length', specs.get('Total length:', ''))
    total_width = specs.get('Total width', specs.get('Total width:', ''))
    max_pass = specs.get('Max Passenger', specs.get('Max Passenger:', ''))
    max_engine = specs.get('Max Engine', specs.get('Max Engine:', ''))
    tube_dia = specs.get('Tube diameter', specs.get('Tube diameter:', ''))
    net_weight = specs.get('Net weight', specs.get('Net weight:', ''))
    air_chamber = specs.get('Air chamber', specs.get('Air chamber:', ''))
    use_case = get_use_case(length_m)

    len_str = str(total_len).strip()
    beam_str = str(total_width).strip().replace('cm', '').strip()
    tube_str = str(tube_dia).strip().replace('cm', '').strip()

    body = f'''# {model_id}

A {len_str} RIB ({model_id}) designed for {use_case} conditions, supporting engines up to {str(max_engine).strip()}.

## Key Features

- **Length:** {len_str}
- **Beam:** {beam_str}cm
- **Capacity:** {max_pass} persons
- **Max Engine:** {str(max_engine).strip()}
- **Air Chambers:** {air_chamber}

The {model_id} features a fiberglass deep-V hull for wave slicing and inflated Hypalon tubes for shock absorption. With a beam of {beam_str}cm, tube diameter of {tube_str}cm, and a weight of {parse_weight(net_weight)}, it supports outboard engines up to {str(max_engine).strip()}.

## Standard Accessories

{chr(10).join(f'- {a}' for a in std_acc) if std_acc else f'The {model_id} includes essential standard equipment for safe and convenient operation.'}

## Optional Accessories

{chr(10).join(f'- {a}' for a in opt_acc) if opt_acc else f'Contact us for available optional accessories and upgrade packages for the {model_id}.'}

## Ideal For

Boaters, fishing enthusiasts, marine professionals, and commercial operators who need a reliable {len_str} RIB for {use_case} operations.

## Compare

Compared to smaller RIB models, the {model_id} offers a {len_str} hull with a beam of {beam_str}cm, making it suitable for {use_case} conditions.
'''
    return body

def main():
    for model_id in sorted(MODEL_NAMES.keys()):
        try:
            specs, std_acc, opt_acc = read_xlsx(model_id)
            frontmatter, slug = generate_frontmatter(model_id, specs, std_acc, opt_acc)

            total_len = specs.get('Total length', specs.get('Total length:', ''))
            length_m, _ = parse_length(total_len)

            body = generate_body(model_id, specs, std_acc, opt_acc, slug, length_m)

            mdx_content = frontmatter + '\n' + body.lstrip('\n')

            out_path = os.path.join(OUT, f'{slug}.mdx')
            with open(out_path, 'w', encoding='utf-8') as f:
                f.write(mdx_content)

            print(f'OK {model_id} -> {slug}.mdx  (air={specs.get("Air chamber", "?")}, war={specs.get("Warranty", specs.get("Warranty: ", "N/A"))}, std_acc={len(std_acc)}, opt_acc={len(opt_acc)})')
        except Exception as e:
            print(f'FAIL {model_id} ERROR: {e}')

if __name__ == '__main__':
    main()
